import openpyxl
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os

# ========== 注册中文字体 ==========
# 尝试多个常见中文字体路径
font_paths = [
    "C:/Windows/Fonts/simhei.ttf",
    "C:/Windows/Fonts/msyh.ttf",
    "C:/Windows/Fonts/simsun.ttc",
    "C:/Windows/Fonts/simfang.ttf",
    "C:/Windows/Fonts/kaiu.ttf",
]

chinese_font = None
for fp in font_paths:
    if os.path.exists(fp):
        try:
            pdfmetrics.registerFont(TTFont('ChineseFont', fp))
            chinese_font = 'ChineseFont'
            print(f"使用字体: {fp}")
            break
        except:
            continue

if not chinese_font:
    # fallback: 尝试搜索
    fonts_dir = "C:/Windows/Fonts"
    for f in os.listdir(fonts_dir):
        if f.lower().endswith(('.ttf', '.ttc')) and any(kw in f.lower() for kw in ['hei', 'ming', 'song', 'kai', 'yahei', 'fang']):
            try:
                pdfmetrics.registerFont(TTFont('ChineseFont', os.path.join(fonts_dir, f)))
                chinese_font = 'ChineseFont'
                print(f"使用字体: {os.path.join(fonts_dir, f)}")
                break
            except:
                continue

if not chinese_font:
    print("警告: 未找到中文字体，将使用默认字体")

FONT_NAME = chinese_font or 'Helvetica'

# ========== 读取 Excel ==========
xlsx_path = r"C:\MAPWARSJ\.hermes\desktop-attachments\Turtle_6-Pattern_4H_ATR_Dynamic_Position_Bull-Bear_Strategy_OKX_BTCUSDT.P_2026-06-13 (4).xlsx"
pdf_path = r"C:\MAPWARSJ\.hermes\desktop-attachments\Turtle_6-Pattern_4H_ATR_Dynamic_Position_Bull-Bear_Strategy_OKX_BTCUSDT.P_2026-06-13 (4).pdf"

wb = openpyxl.load_workbook(xlsx_path, data_only=True)

# ========== 生成 PDF ==========
# 对大表格使用横版
page_w, page_h = landscape(A4)  # 横版

doc = SimpleDocTemplate(
    pdf_path,
    pagesize=(page_w, page_h),
    leftMargin=10*mm,
    rightMargin=10*mm,
    topMargin=10*mm,
    bottomMargin=10*mm,
)

styles = getSampleStyleSheet()
title_style = ParagraphStyle(
    'ChineseTitle',
    parent=styles['Heading1'],
    fontName=FONT_NAME,
    fontSize=14,
    leading=18,
    spaceAfter=6,
)
normal_style = ParagraphStyle(
    'ChineseNormal',
    parent=styles['Normal'],
    fontName=FONT_NAME,
    fontSize=7,
    leading=10,
)

elements = []

def cell_value(v):
    """格式化单元格"""
    if v is None:
        return ''
    if isinstance(v, float):
        return f"{v:.4f}" if abs(v) < 1000 else f"{v:.2f}"
    return str(v)

def sheet_to_table(ws, max_width_mm=277):
    """将工作表转为 ReportLab Table，自动缩放"""
    rows_data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows_data.append([cell_value(v) for v in row])

    if not rows_data:
        return None

    # 计算列宽：根据内容
    col_count = len(rows_data[0])
    row_count = len(rows_data)

    # 估算每列最大宽度
    col_widths_chars = [0] * col_count
    for row in rows_data:
        for i, cell in enumerate(row):
            if i < col_count:
                col_widths_chars[i] = max(col_widths_chars[i], len(str(cell)))

    # 每个字符约 5pt，但要设上下限
    available_mm = max_width_mm
    total_chars = sum(col_widths_chars)
    if total_chars > 0:
        scale = available_mm / (total_chars * 1.8)  # 1.8mm per char approx
        col_widths_mm = [max(10, min(60, c * 1.8 * scale)) for c in col_widths_chars]
    else:
        col_widths_mm = [available_mm / col_count] * col_count

    # 确保总和不超过可用宽度
    total = sum(col_widths_mm)
    if total > available_mm:
        ratio = available_mm / total
        col_widths_mm = [w * ratio for w in col_widths_mm]

    table = Table(rows_data, colWidths=[w * mm for w in col_widths_mm], repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
        ('FONTSIZE', (0, 0), (-1, -1), 6.5),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    return table

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"处理工作表: {sname} ({ws.max_row} 行 × {ws.max_column} 列)")

    elements.append(Paragraph(f"📊 {sname}", title_style))
    elements.append(Spacer(1, 4*mm))

    tbl = sheet_to_table(ws)
    if tbl:
        elements.append(tbl)

    elements.append(PageBreak())

# 移除最后一个多余的 PageBreak
if elements and isinstance(elements[-1], PageBreak):
    elements.pop()

print(f"\n生成 PDF: {pdf_path}")
doc.build(elements)

size_kb = os.path.getsize(pdf_path) / 1024
print(f"完成! 文件大小: {size_kb:.1f} KB")
